"""
Generate Excel Report for Last 6 Months
Columns: Date, Instrument Name, Entry Price, Exit Price, Profit, Trade Strategy, Cumulative P&L
"""

from dhanhq import dhanhq, DhanContext
from config import CLIENT_ID, ACCESS_TOKEN
import strategy_config as cfg
import pandas as pd
from datetime import datetime, timedelta
import numpy as np


class ExcelReportGenerator:
    def __init__(self):
        dhan_context = DhanContext(CLIENT_ID, ACCESS_TOKEN)
        self.dhan = dhanhq(dhan_context)
        self.trades = []
        
        self.max_loss = cfg.MAX_LOSS_PER_LOT
        self.target = cfg.TARGET_PER_LOT
        self.lot_size = cfg.LOT_SIZE
        self.itm_points = cfg.ITM_POINTS
        self.trades_per_day = cfg.TRADES_PER_DAY
        
    def get_historical_data(self, days=180):
        """Get NIFTY historical data for 6 months"""
        try:
            to_date = datetime.now().strftime('%Y-%m-%d')
            from_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
            
            print(f"📊 Fetching 6 months data ({days} days)...")
            print(f"   From: {from_date} to {to_date}")
            
            data = self.dhan.historical_daily_data(
                security_id="13",
                exchange_segment="NSE_EQ",
                instrument_type="INDEX",
                from_date=from_date,
                to_date=to_date
            )
            
            if data and 'data' in data:
                df = pd.DataFrame(data['data'])
                
                column_mapping = {
                    'open': 'Open',
                    'high': 'High',
                    'low': 'Low',
                    'close': 'Close',
                    'volume': 'Volume'
                }
                
                for old_col, new_col in column_mapping.items():
                    if old_col in df.columns:
                        df.rename(columns={old_col: new_col}, inplace=True)
                
                if 'timestamp' in df.columns:
                    if pd.api.types.is_numeric_dtype(df['timestamp']):
                        df['Date'] = pd.to_datetime(df['timestamp'], unit='s')
                    else:
                        df['Date'] = pd.to_datetime(df['timestamp'])
                
                df = df.sort_values('Date').reset_index(drop=True)
                
                # Add technical indicators
                df['EMA20'] = df['Close'].ewm(span=20).mean()
                df['EMA50'] = df['Close'].ewm(span=50).mean()
                df['SMA20'] = df['Close'].rolling(20).mean()
                
                print(f"✅ Loaded {len(df)} days of data")
                print(f"   Date range: {df['Date'].min().date()} to {df['Date'].max().date()}\n")
                
                return df
            
            return None
            
        except Exception as e:
            print(f"❌ Error: {e}")
            return None
    
    def calculate_fibonacci_levels(self, high, low):
        """Calculate Fibonacci levels"""
        diff = high - low
        return {
            '0.0': high,
            '0.236': high - (diff * 0.236),
            '0.382': high - (diff * 0.382),
            '0.500': high - (diff * 0.500),
            '0.618': high - (diff * 0.618),
            '0.786': high - (diff * 0.786),
            '1.0': low
        }
    
    def detect_all_signals(self, df, idx):
        """Detect all trading signals"""
        if idx + 1 >= len(df) or idx < 20:
            return []
        
        signals = []
        
        # Fibonacci
        fib_signal = self.detect_fibonacci(df, idx)
        if fib_signal:
            signals.append(('Fibonacci', fib_signal))
        
        # Candlestick
        candle_signal = self.detect_candlestick(df, idx)
        if candle_signal:
            signals.append(('Candlestick', candle_signal))
        
        # EMA Bounce
        ema_signal = self.detect_ema_bounce(df, idx)
        if ema_signal:
            signals.append(('EMA Bounce', ema_signal))
        
        # S/R
        sr_signal = self.detect_support_resistance(df, idx)
        if sr_signal:
            signals.append(('Support/Resistance', sr_signal))
        
        return signals
    
    def detect_fibonacci(self, df, idx):
        """Detect Fibonacci setup"""
        lookback = 10
        if idx < lookback:
            return None
        
        current = df.iloc[idx]
        lookback_data = df.iloc[idx-lookback:idx]
        
        swing_high = lookback_data['High'].max()
        swing_low = lookback_data['Low'].min()
        
        fib_levels = self.calculate_fibonacci_levels(swing_high, swing_low)
        tolerance = 150
        
        if abs(current['Low'] - fib_levels['0.618']) < tolerance and current['Close'] > current['Open']:
            if current['Close'] > current['SMA20']:
                return {'type': 'CALL', 'signal': 'Fib 61.8% Bounce'}
        
        if abs(current['High'] - fib_levels['0.382']) < tolerance and current['Close'] < current['Open']:
            if current['Close'] < current['SMA20']:
                return {'type': 'PUT', 'signal': 'Fib 38.2% Rejection'}
        
        return None
    
    def detect_candlestick(self, df, idx):
        """Detect candlestick patterns"""
        if idx < 1:
            return None
        
        current = df.iloc[idx]
        o, h, l, c = current['Open'], current['High'], current['Low'], current['Close']
        body = abs(c - o)
        upper_shadow = h - max(o, c)
        lower_shadow = min(o, c) - l
        
        if lower_shadow > body * 2 and upper_shadow < body * 0.5 and c > o:
            if current['Close'] > current['EMA20']:
                return {'type': 'CALL', 'signal': 'Hammer'}
        
        if upper_shadow > body * 2 and lower_shadow < body * 0.5 and c < o:
            if current['Close'] < current['EMA20']:
                return {'type': 'PUT', 'signal': 'Shooting Star'}
        
        return None
    
    def detect_ema_bounce(self, df, idx):
        """Detect EMA bounce"""
        if idx < 2:
            return None
        
        current = df.iloc[idx]
        prev = df.iloc[idx-1]
        ema20 = current['EMA20']
        
        if prev['Low'] <= ema20 * 1.005 and current['Close'] > ema20:
            if current['Close'] > current['EMA50']:
                return {'type': 'CALL', 'signal': '20 EMA Bounce'}
        
        if prev['High'] >= ema20 * 0.995 and current['Close'] < ema20:
            if current['Close'] < current['EMA50']:
                return {'type': 'PUT', 'signal': '20 EMA Rejection'}
        
        return None
    
    def detect_support_resistance(self, df, idx):
        """Detect S/R bounce"""
        lookback = 10
        if idx < lookback:
            return None
        
        current = df.iloc[idx]
        lookback_data = df.iloc[idx-lookback:idx]
        
        resistance = lookback_data['High'].max()
        support = lookback_data['Low'].min()
        tolerance = 50
        
        if abs(current['Low'] - support) < tolerance and current['Close'] > current['Open']:
            if current['Close'] > current['SMA20']:
                return {'type': 'CALL', 'signal': 'Support Bounce'}
        
        if abs(current['High'] - resistance) < tolerance and current['Close'] < current['Open']:
            if current['Close'] < current['SMA20']:
                return {'type': 'PUT', 'signal': 'Resistance Rejection'}
        
        return None
    
    def simulate_option_price(self, spot_price, strike, option_type):
        """Simplified option pricing"""
        if option_type == "CE":
            intrinsic = max(0, spot_price - strike)
        else:
            intrinsic = max(0, strike - spot_price)
        
        time_value = spot_price * 0.015
        return max(intrinsic + time_value, 30)
    
    def check_trade_outcome(self, entry_price, next_day_high, next_day_low, next_day_close, 
                           entry_premium, option_type, strike):
        """
        Advanced trailing SL logic:
        1. If target hit -> Move SL to entry (lock profit)
        2. After target -> Use 10-point trailing SL for additional gains
        3. If SL hit after target -> Exit at target level (protected profit)
        """
        sl_per_contract = self.max_loss / self.lot_size
        target_per_contract = self.target / self.lot_size

        target_premium = entry_premium + target_per_contract
        sl_premium = entry_premium - sl_per_contract

        high_premium = self.simulate_option_price(next_day_high, strike, option_type)
        low_premium = self.simulate_option_price(next_day_low, strike, option_type)
        close_premium = self.simulate_option_price(next_day_close, strike, option_type)

        # Check if target is hit
        if high_premium >= target_premium:
            # Target hit! Now check for additional gains with tighter trailing

            # Move SL to entry price (lock in target profit)
            new_sl_premium = entry_premium

            # Calculate potential additional gain with 10-point trailing
            additional_gain_per_point = self.lot_size / 50  # Premium per index point
            trailing_sl_after_target = 10  # Tighter trailing after target

            # Maximum premium reached during the day
            max_premium_after_target = high_premium

            # Trailing SL level (10 points below max)
            trailing_sl_premium = max_premium_after_target - (trailing_sl_after_target * additional_gain_per_point)

            # If close is above trailing SL, we get additional profit
            if close_premium >= trailing_sl_premium:
                # Held position with trailing SL - got additional profit
                pnl = (close_premium - entry_premium) * self.lot_size
                return close_premium, pnl, "Target + Trailing"
            else:
                # Trailing SL hit, but we're still at breakeven or better
                # Exit at the trailing SL level (but minimum at entry to lock target profit)
                exit_premium = max(trailing_sl_premium, entry_premium)
                pnl = (exit_premium - entry_premium) * self.lot_size
                return exit_premium, pnl, "Target + TSL Hit"

        # Target not hit - check initial SL
        if low_premium <= sl_premium:
            return sl_premium, -self.max_loss, "SL Hit"

        # Neither target nor SL hit - close at market close
        pnl = (close_premium - entry_premium) * self.lot_size
        return close_premium, pnl, "Close Exit"

    def backtest_day(self, df, idx):
        """Get up to 2 trades per day"""
        
        all_signals = self.detect_all_signals(df, idx)
        
        if not all_signals:
            return []
        
        current = df.iloc[idx]
        next_day = df.iloc[idx + 1]
        date = current['Date']
        
        trades = []
        
        for i, (strategy_name, signal_dict) in enumerate(all_signals[:self.trades_per_day]):
            option_type = signal_dict['type']
            signal_name = signal_dict['signal']
            
            entry_price = current['Close']
            
            if option_type == 'CALL':
                strike = round((entry_price - self.itm_points) / 50) * 50
                instrument_name = f"NIFTY {strike} CE"
            else:
                strike = round((entry_price + self.itm_points) / 50) * 50
                instrument_name = f"NIFTY {strike} PE"
            
            entry_premium = self.simulate_option_price(entry_price, strike, "CE" if option_type == "CALL" else "PE")
            
            exit_premium, pnl, exit_reason = self.check_trade_outcome(
                entry_price, next_day['High'], next_day['Low'], next_day['Close'],
                entry_premium, "CE" if option_type == "CALL" else "PE", strike
            )
            
            trade = {
                'Date': date,
                'Instrument Name': instrument_name,
                'Entry Price': round(entry_premium, 2),
                'Exit Price': round(exit_premium, 2),
                'Profit': round(pnl, 2),
                'Trade Strategy': strategy_name,
                'Signal': signal_name,
                'Exit Reason': exit_reason
            }
            
            trades.append(trade)
        
        return trades
    
    def run_backtest(self, months=6):
        """Run backtest for 6 months"""
        
        trading_days = months * 20  # Approx 20 trading days per month
        
        print("=" * 80)
        print(f"📊 GENERATING EXCEL REPORT FOR LAST {months} MONTHS")
        print("=" * 80)
        print(f"Expected Trading Days: {trading_days}")
        print(f"Max Trades Per Day: {self.trades_per_day}")
        print("=" * 80)
        print()
        
        df = self.get_historical_data(days=months*30)
        
        if df is None:
            return None
        
        print(f"🔍 Scanning for trades...\n")
        
        self.trades = []
        days_traded = 0
        
        for idx in range(20, len(df) - 1):
            day_trades = self.backtest_day(df, idx)
            
            if day_trades:
                days_traded += 1
                self.trades.extend(day_trades)
                print(f"Day {days_traded}: {day_trades[0]['Date'].date()} - {len(day_trades)} trade(s)")
            
            if days_traded >= trading_days:
                break
        
        return self.trades
    
    def export_to_excel(self, filename='6_MONTHS_TRADING_REPORT.xlsx'):
        """Export trades to Excel with proper formatting"""
        
        if not self.trades:
            print("❌ No trades to export")
            return
        
        print("\n📊 Generating Excel Report...\n")
        
        # Create DataFrame
        df = pd.DataFrame(self.trades)
        
        # Add cumulative P&L column
        df['Cumulative P&L'] = df['Profit'].cumsum().round(2)
        
        # Format date column
        df['Date'] = pd.to_datetime(df['Date']).dt.strftime('%Y-%m-%d')
        
        # Select and reorder columns for export
        export_columns = [
            'Date',
            'Instrument Name',
            'Entry Price',
            'Exit Price',
            'Profit',
            'Trade Strategy',
            'Cumulative P&L'
        ]
        
        df_export = df[export_columns].copy()
        
        # Create Excel writer
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Write main data
            df_export.to_excel(writer, sheet_name='Trading Report', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Trading Report']
            
            # Format headers
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=12)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format numbers
            for row in worksheet.iter_rows(min_row=2, max_row=len(df_export)+1):
                # Entry Price, Exit Price
                for cell in [row[2], row[3]]:
                    cell.number_format = '₹#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                
                # Profit
                cell = row[4]
                cell.number_format = '₹#,##0.00'
                cell.alignment = Alignment(horizontal='right')
                if cell.value < 0:
                    cell.font = Font(color='FF0000')  # Red for losses
                else:
                    cell.font = Font(color='008000')  # Green for profits
                
                # Cumulative P&L
                cell = row[6]
                cell.number_format = '₹#,##0.00'
                cell.alignment = Alignment(horizontal='right')
                cell.font = Font(bold=True)
            
            # Add summary row
            summary_row = len(df_export) + 3
            worksheet[f'A{summary_row}'] = 'TOTAL'
            worksheet[f'A{summary_row}'].font = Font(bold=True, size=12)
            
            total_profit = df['Profit'].sum()
            worksheet[f'E{summary_row}'] = total_profit
            worksheet[f'E{summary_row}'].number_format = '₹#,##0.00'
            worksheet[f'E{summary_row}'].font = Font(bold=True, size=12)
            worksheet[f'E{summary_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            # Add statistics sheet
            stats_data = {
                'Metric': [
                    'Total Trades',
                    'Winning Trades',
                    'Losing Trades',
                    'Win Rate %',
                    'Total Profit',
                    'Total Loss',
                    'Net P&L',
                    'Profit Factor',
                    'Average Profit per Trade',
                    'Largest Win',
                    'Largest Loss'
                ],
                'Value': [
                    len(df),
                    len(df[df['Profit'] > 0]),
                    len(df[df['Profit'] <= 0]),
                    f"{(len(df[df['Profit'] > 0]) / len(df) * 100):.2f}%",
                    f"Rs.{df[df['Profit'] > 0]['Profit'].sum():,.2f}",
                    f"Rs.{df[df['Profit'] <= 0]['Profit'].sum():,.2f}",
                    f"Rs.{df['Profit'].sum():,.2f}",
                    f"{df[df['Profit'] > 0]['Profit'].sum() / abs(df[df['Profit'] <= 0]['Profit'].sum()):.2f}" if df[df['Profit'] <= 0]['Profit'].sum() != 0 else "N/A",
                    f"Rs.{df['Profit'].mean():,.2f}",
                    f"Rs.{df['Profit'].max():,.2f}",
                    f"Rs.{df['Profit'].min():,.2f}"
                ]
            }
            
            df_stats = pd.DataFrame(stats_data)
            df_stats.to_excel(writer, sheet_name='Statistics', index=False)
            
            # Format statistics sheet
            stats_sheet = writer.sheets['Statistics']
            for cell in stats_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            stats_sheet.column_dimensions['A'].width = 30
            stats_sheet.column_dimensions['B'].width = 25
        
        print(f"✅ Excel report generated: {filename}")
        print(f"   Total Trades: {len(df)}")
        print(f"   Total P&L: Rs.{df['Profit'].sum():,.2f}")
        print(f"   Win Rate: {(len(df[df['Profit'] > 0]) / len(df) * 100):.1f}%")


if __name__ == "__main__":
    import sys
    import io
    if sys.stdout.encoding != 'utf-8':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    
    print("\n🚀 Generating 6-Month Trading Report...\n")
    
    generator = ExcelReportGenerator()
    trades = generator.run_backtest(months=6)
    
    if trades:
        generator.export_to_excel('6_MONTHS_TRADING_REPORT.xlsx')
        print("\n✅ Report generation complete!")
        print("\n📁 File saved: 6_MONTHS_TRADING_REPORT.xlsx")
        print("   Open this file in Microsoft Excel")
    else:
        print("\n⚠️  No trades found")
